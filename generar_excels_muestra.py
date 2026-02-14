#!/usr/bin/env python3
"""Genera Excel de muestra para importar en el Sistema de Comisiones."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), 'excel_muestra')
os.makedirs(OUTPUT_DIR, exist_ok=True)

HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
ABONO_FILL = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

def style_data(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 35)

# ==============================
# 1. SHOWROOMS
# ==============================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Showrooms'
ws.append(['Nombre', 'Comisión (%)'])
showrooms = [
    ('Milano Fashion Hub', 15),
    ('Paris Luxe Showroom', 12),
    ('London Style Room', 18),
    ('NYC Premium Space', 20),
    ('Berlin Mode Gallery', 14),
]
for s in showrooms:
    ws.append(s)
style_header(ws)
style_data(ws)
auto_width(ws)
wb.save(os.path.join(OUTPUT_DIR, 'Showrooms_muestra.xlsx'))
print('✓ Showrooms_muestra.xlsx')

# ==============================
# 2. CLIENTES
# ==============================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Clientes'
ws.append(['Nombre', 'Showroom'])
clientes = [
    ('Boutique Rossi', 'Milano Fashion Hub'),
    ('Maison Dupont', 'Paris Luxe Showroom'),
    ('Casa Bianchi', 'Milano Fashion Hub'),
    ('Harvey & Sons', 'London Style Room'),
    ('Bergdorf Select', 'NYC Premium Space'),
    ('Atelier Laurent', 'Paris Luxe Showroom'),
    ('KaDeWe Luxury', 'Berlin Mode Gallery'),
    ('Selfridges Concession', 'London Style Room'),
    ('Saks Fifth Avenue', 'NYC Premium Space'),
    ('Galeries Lafayette', 'Paris Luxe Showroom'),
]
for c in clientes:
    ws.append(c)
style_header(ws)
style_data(ws)
auto_width(ws)
wb.save(os.path.join(OUTPUT_DIR, 'Clientes_muestra.xlsx'))
print('✓ Clientes_muestra.xlsx')

# ==============================
# 3. PEDIDOS
# ==============================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Pedidos'
ws.append(['Número', 'Cliente', 'Fecha', 'Moneda', 'Importe'])
pedidos = [
    ('PED-2025-001', 'Boutique Rossi',        '2025-01-15', 'EUR', 12500.00),
    ('PED-2025-002', 'Boutique Rossi',        '2025-02-10', 'EUR', 8750.00),
    ('PED-2025-003', 'Maison Dupont',         '2025-01-20', 'EUR', 15300.00),
    ('PED-2025-004', 'Casa Bianchi',          '2025-02-01', 'EUR', 6200.00),
    ('PED-2025-005', 'Harvey & Sons',         '2025-01-25', 'EUR', 9800.00),
    ('PED-2025-006', 'Harvey & Sons',         '2025-03-05', 'EUR', 4500.00),
    ('PED-2025-007', 'Bergdorf Select',       '2025-02-15', 'USD', 22000.00),
    ('PED-2025-008', 'Bergdorf Select',       '2025-03-01', 'USD', 18500.00),
    ('PED-2025-009', 'Atelier Laurent',       '2025-01-10', 'EUR', 11200.00),
    ('PED-2025-010', 'KaDeWe Luxury',         '2025-02-20', 'EUR', 7600.00),
    ('PED-2025-011', 'Selfridges Concession', '2025-03-10', 'EUR', 13400.00),
    ('PED-2025-012', 'Saks Fifth Avenue',     '2025-02-28', 'USD', 16800.00),
    ('PED-2025-013', 'Galeries Lafayette',    '2025-03-15', 'EUR', 19500.00),
    ('PED-2025-014', 'Boutique Rossi',        '2025-04-01', 'EUR', 5300.00),
    ('PED-2025-015', 'Maison Dupont',         '2025-04-10', 'EUR', 7800.00),
]
for p in pedidos:
    ws.append(p)
style_header(ws)
style_data(ws)
auto_width(ws)
wb.save(os.path.join(OUTPUT_DIR, 'Pedidos_muestra.xlsx'))
print('✓ Pedidos_muestra.xlsx')

# ==============================
# 4. FACTURAS (con abonos)
# ==============================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Facturas'
ws.append(['Número', 'Cliente', 'Pedidos', 'Fecha', 'Vencimiento', 'Moneda', 'Importe', 'FacturasAbonadas'])

facturas = [
    # Facturas normales vinculadas a pedidos
    ('FAC-2025-001', 'Boutique Rossi',        'PED-2025-001',            '2025-02-01', '2025-04-01', 'EUR', 12500.00, ''),
    ('FAC-2025-002', 'Boutique Rossi',        'PED-2025-002',            '2025-03-01', '2025-05-01', 'EUR', 8750.00,  ''),
    ('FAC-2025-003', 'Maison Dupont',         'PED-2025-003',            '2025-02-15', '2025-04-15', 'EUR', 15300.00, ''),
    ('FAC-2025-004', 'Casa Bianchi',          'PED-2025-004',            '2025-03-01', '2025-05-01', 'EUR', 6200.00,  ''),
    ('FAC-2025-005', 'Harvey & Sons',         'PED-2025-005',            '2025-02-20', '2025-04-20', 'EUR', 9800.00,  ''),
    ('FAC-2025-006', 'Harvey & Sons',         'PED-2025-006',            '2025-04-01', '2025-06-01', 'EUR', 4500.00,  ''),
    ('FAC-2025-007', 'Bergdorf Select',       'PED-2025-007',            '2025-03-15', '2025-05-15', 'USD', 22000.00, ''),
    ('FAC-2025-008', 'Bergdorf Select',       'PED-2025-008',            '2025-04-01', '2025-06-01', 'USD', 18500.00, ''),
    ('FAC-2025-009', 'Atelier Laurent',       'PED-2025-009',            '2025-02-10', '2025-04-10', 'EUR', 11200.00, ''),
    ('FAC-2025-010', 'KaDeWe Luxury',         'PED-2025-010',            '2025-03-15', '2025-05-15', 'EUR', 7600.00,  ''),
    ('FAC-2025-011', 'Selfridges Concession', 'PED-2025-011',            '2025-04-10', '2025-06-10', 'EUR', 13400.00, ''),
    ('FAC-2025-012', 'Saks Fifth Avenue',     'PED-2025-012',            '2025-03-25', '2025-05-25', 'USD', 16800.00, ''),
    ('FAC-2025-013', 'Galeries Lafayette',    'PED-2025-013',            '2025-04-15', '2025-06-15', 'EUR', 19500.00, ''),
    ('FAC-2025-014', 'Boutique Rossi',        'PED-2025-014',            '2025-05-01', '2025-07-01', 'EUR', 5300.00,  ''),
    ('FAC-2025-015', 'Maison Dupont',         'PED-2025-015',            '2025-05-10', '2025-07-10', 'EUR', 7800.00,  ''),
    # Factura con múltiples pedidos
    ('FAC-2025-016', 'Harvey & Sons',         'PED-2025-005, PED-2025-006', '2025-04-20', '2025-06-20', 'EUR', 14300.00, ''),

    # --- ABONOS / RECTIFICATIVAS ---
    # Abono parcial de FAC-2025-001 (factura de 12.500€ -> abono de 2.000€)
    ('ABN-2025-001', 'Boutique Rossi',        '',                        '2025-03-15', '2025-03-15', 'EUR', 2000.00,  'FAC-2025-001'),
    # Abono total de FAC-2025-004 (factura de 6.200€ -> devolución completa)
    ('ABN-2025-002', 'Casa Bianchi',          '',                        '2025-04-01', '2025-04-01', 'EUR', 6200.00,  'FAC-2025-004'),
    # Abono parcial en USD de FAC-2025-007 (factura de $22.000 -> abono de $3.500)
    ('ABN-2025-003', 'Bergdorf Select',       '',                        '2025-04-15', '2025-04-15', 'USD', 3500.00,  'FAC-2025-007'),
    # Abono sobre factura ya cobrada (escenario 3: factura ya saldada, abono independiente)
    ('ABN-2025-004', 'Maison Dupont',         '',                        '2025-05-20', '2025-05-20', 'EUR', 1500.00,  'FAC-2025-003'),
    # Abono que referencia dos facturas
    ('ABN-2025-005', 'Harvey & Sons',         '',                        '2025-05-01', '2025-05-01', 'EUR', 800.00,   'FAC-2025-005, FAC-2025-006'),
]

for idx, f in enumerate(facturas):
    ws.append(f)
    # Resaltar filas de abono
    if f[7]:  # tiene FacturasAbonadas -> es abono
        row_num = idx + 2  # +1 header, +1 zero-based
        for cell in ws[row_num]:
            cell.fill = ABONO_FILL

style_header(ws)
style_data(ws)
auto_width(ws)

# Añadir nota explicativa
ws.append([])
ws.append(['NOTAS:'])
ws.append(['- Las filas en rojo son abonos/rectificativas. El importe se puede poner en positivo; el sistema lo convierte a negativo.'])
ws.append(['- La columna "FacturasAbonadas" indica qué factura(s) se abonan. Se pueden separar con comas.'])
ws.append(['- La columna "Pedidos" no aplica a abonos (se ignora si hay FacturasAbonadas).'])
ws.append([f'- Formato: Número | Cliente | Pedidos | Fecha | Vencimiento | Moneda | Importe | FacturasAbonadas'])

note_font = Font(italic=True, color='666666', size=10)
for row in ws.iter_rows(min_row=ws.max_row - 4, max_row=ws.max_row, max_col=1):
    for cell in row:
        cell.font = note_font

wb.save(os.path.join(OUTPUT_DIR, 'Facturas_con_abonos_muestra.xlsx'))
print('✓ Facturas_con_abonos_muestra.xlsx')

print(f'\nArchivos generados en: {OUTPUT_DIR}/')
